import random
from openpyxl import Workbook, load_workbook

def construirTabuleiro(qtd_linhas, qtd_colunas, qtd_bombas):
 
    arquivo_excel = Workbook()

    tabuleiro = arquivo_excel.active

    # Lançar bombas aleatoriamente
    conta_bombas = 1
    while(conta_bombas <= qtd_bombas):
        linha_aleatoria = random.randint(1,qtd_linhas)
        coluna_aleatoria = random.randint(1,qtd_colunas)

        if tabuleiro.cell(row=linha_aleatoria, column=coluna_aleatoria).value != '*':
            tabuleiro.cell(row=linha_aleatoria, column=coluna_aleatoria, value='*')
            conta_bombas+=1

    # Preencher números
    
    for linha in range(1, qtd_linhas+1):
        for coluna in range(1, qtd_colunas+1):
            if tabuleiro.cell(row=linha, column=coluna).value != '*':
                bombas_aoredor = 0
                if (coluna+1)<=qtd_colunas and tabuleiro.cell(row=linha, column=coluna+1).value == '*':
                    bombas_aoredor+=1
                if (linha+1)<=qtd_linhas and (coluna+1)<=qtd_colunas and tabuleiro.cell(row=linha+1, column=coluna+1).value == '*':
                    bombas_aoredor+=1
                if (linha+1)<=qtd_linhas and tabuleiro.cell(row=linha+1, column=coluna).value == '*':
                    bombas_aoredor+=1
                if (linha+1)<=qtd_linhas and (coluna-1)>=1 and tabuleiro.cell(row=linha+1, column=coluna-1).value == '*':
                    bombas_aoredor+=1
                if (coluna-1)>=1 and tabuleiro.cell(row=linha, column=coluna-1).value == '*':
                    bombas_aoredor+=1
                if (coluna-1)>=1 and (linha-1)>=1 and tabuleiro.cell(row=linha-1, column=coluna-1).value == '*':
                    bombas_aoredor+=1
                if (linha-1)>=1 and tabuleiro.cell(row=linha-1, column=coluna).value == '*':
                    bombas_aoredor+=1
                if (coluna+1)<=qtd_colunas and (linha-1)>=1 and tabuleiro.cell(row=linha-1, column=coluna+1).value == '*':
                    bombas_aoredor+=1
                tabuleiro.cell(row=linha, column=coluna, value=bombas_aoredor)
    arquivo_excel.save('tabuleiro.xlsx')

def iniciar():

    print('\n\n*** Bem-vindo ao Jogo Campo Minado ***\n\n')

    print('Qual é o tamanho do tabuleiro que você deseja jogar?')

    qtd_linhas = int(input('Quantidade de Linhas = ')) 
    qtd_colunas = int(input('Quantidade de Colunas = '))

    print('\nEscolha o nível do jogo: ')
    print('1 - FACIL')
    print('2 - MÉDIO')
    print('3 - DIFÍCIL')

    nivel = int(input('Qual é o nível que você deseja jogar?'))
    qtd_bombas = 0

    if nivel == 1:
        qtd_bombas = (qtd_linhas * qtd_colunas) * 0.15
    if nivel == 2:
        qtd_bombas = (qtd_linhas * qtd_colunas) * 0.40
    if nivel == 3:
        qtd_bombas = (qtd_linhas * qtd_colunas) * 0.60

    qtd_bombas = int(qtd_bombas)

    construirTabuleiro(qtd_linhas, qtd_colunas, qtd_bombas)

def mostrarTabuleiro(tabuleiro):
    print('   ', end='')

    for coluna in range(1, tabuleiro.max_column+1):
        print('%3i' % coluna, end='')

    print('')

    for linha in range(1, tabuleiro.max_row+1):
        print('%3i' % linha, end='')
        for coluna in range(1, tabuleiro.max_column+1):
            print('%3c' % tabuleiro.cell(row=linha, column=coluna).value, end='')
        print('')

def jogar():
    arquivo_excel = load_workbook('tabuleiro.xlsx', read_only = False)

    tab_oculto = arquivo_excel.active
    tab_oculto.title = 'tabuleiro_oculto'

    tab_usuario = arquivo_excel.create_sheet('tabuleiro_usuario')

    # Preencher '-' no tabuleiro usuário  
    for linha in range(1, tab_oculto.max_row+1):
        for coluna in range(1, tab_oculto.max_column+1):
            tab_usuario.cell(row=linha, column=coluna, value='-')

    mostrarTabuleiro(tab_usuario)
    
    arquivo_excel.save('tabuleiro.xlsx')
    
# Início do Jogo

iniciar()

jogar()


