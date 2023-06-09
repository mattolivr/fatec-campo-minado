import os
import random
import math

from openpyxl import Workbook

class Tabuleiro():
    def __init__(self, altura: int, largura: int, dificuldade: float):
        self.altura = altura
        self.largura = largura
        self.dificuldade = dificuldade

        self.campo = list()
        self.bombas = list()
        self.excel = Workbook()

        self.__montaTabuleiro()
        self.__sorteiaBombas()
        self.__aplicaBombas(self.bombas)
        self.__montaExcel()

    def toString(self):
        # TODO: Tratar valores com 2 casas decimais
        tabuleiroString = "  "
        for coluna in range(self.largura):
            tabuleiroString += str(coluna + 1) + " "
        tabuleiroString += os.linesep

        for linha in range(self.altura):
            colunas = " ".join(map(str, self.campo[linha]))
            tabuleiroString += str(linha + 1) + " " + colunas

            if (linha != self.altura):
                tabuleiroString += os.linesep
        return tabuleiroString

    def __aplicaBombas(self, bombas: list[list[int, int]]):
        for bomba in bombas:
            linha = bomba[0] - 1
            coluna = bomba[1] - 1

            self.campo[linha][coluna] = 9

            for i in range(-1, 2):
                for j in range(-1, 2):
                    nova_linha = linha + i
                    nova_coluna = coluna + j

                    if (0 <= nova_linha < self.altura) and (0 <= nova_coluna < self.largura) and (self.campo[nova_linha][nova_coluna] != 9):
                        self.campo[nova_linha][nova_coluna] += 1

    def __montaTabuleiro(self):
        for linha in range(self.altura):
            tabuleiroLinha = list()
            for coluna in range(self.largura):
                tabuleiroLinha.append(0)
            self.campo.append(tabuleiroLinha)

    def __sorteiaBombas(self):
        quantidadeBombas = math.ceil((self.altura * self.largura) * self.dificuldade)
        bombasInseridas = 0

        while(bombasInseridas < quantidadeBombas):
            novaBomba = [random.randint(1, self.altura), random.randint(1, self.altura)]

            if (not self.__existeBomba(novaBomba)):
                self.bombas.append(novaBomba)
                bombasInseridas += 1

    def __existeBomba(self, novaBomba: list[int]):
        for bomba in self.bombas:
            if (bomba[0] == novaBomba[0] and bomba[1] == novaBomba[1]):
                return True
        return False       


    def __montaExcel(self):
        planilha = self.excel.active
        planilha.title = "tabuleiro"

        for row in planilha.iter_rows():
            for cell in row:
                cell.value = None

        for linha in range(self.altura):
            for coluna in range(self.largura):
                planilha.cell(row = linha + 1, column = coluna + 1, value = self.campo[linha][coluna])

        self.excel.save("tabuleiro.xlsx")

    def __getValorExcel(self, linha: int, coluna: int):
        planilha = self.excel.active

        return planilha.cell(row = linha + 1, column = coluna + 1).value

